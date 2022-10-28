import eel
import openpyxl

wb = openpyxl.load_workbook("tasks.xlsx")
sheets = wb.sheetnames
main_sheet = wb.active


@eel.expose
def show_tasks():
    task_list = list()
    for i in range(2, main_sheet.max_row + 1):
        task_list.append(main_sheet[f'A{i}'].value)

    return task_list


@eel.expose
def add_task_py(task):
    last_row = int(main_sheet.max_row + 1)
    main_sheet[f"A{last_row}"] = task
    wb.save("tasks.xlsx")


@eel.expose
def delete_task(task):
    for i in range(2, main_sheet.max_row + 1):
        if(main_sheet[f'A{i}'].value == task):
            main_sheet.delete_rows(idx=i, amount=1)
    wb.save("tasks.xlsx")


@eel.expose 
def delete_all_tasks():
    main_sheet.delete_rows(idx=2, amount=main_sheet.max_row)
    wb.save("tasks.xlsx")


def main():
    eel.init("web")
    eel.start("index.html", size=(700, 700))


if __name__ == "__main__":
    main()
